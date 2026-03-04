from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from typing import List


class ExcelExporter:
    def export(self, tables: List[List[List[str]]], output_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Table 1"

        merged_rows = self._merge_tables(tables)
        if merged_rows:
            max_cols = max(len(row) for row in merged_rows)
            for r_idx, row in enumerate(merged_rows, 1):
                padded = list(row) + [""] * (max_cols - len(row))
                for c_idx, value in enumerate(padded, 1):
                    cell = ws.cell(r_idx, c_idx, value=value)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if r_idx == 1:
                        cell.fill = PatternFill("solid", fgColor="2E86AB")
                        cell.font = Font(bold=True, color="FFFFFF")

            for c_idx in range(1, max_cols + 1):
                letter = get_column_letter(c_idx)
                max_len = max((len(str(ws.cell(r, c_idx).value or "")) for r in range(1, ws.max_row + 1)))
                ws.column_dimensions[letter].width = min(max_len + 4, 60)

            ws.freeze_panes = "A2"

        wb.save(output_path)

    @staticmethod
    def _merge_tables(tables: List[List[List[str]]]) -> List[List[str]]:
        merged: List[List[str]] = []
        header: List[str] | None = None

        for table in tables:
            if not table:
                continue

            if header is None:
                header = table[0]
                merged.extend(table)
                continue

            table_rows = table
            if table and ExcelExporter._normalize_row(table[0]) == ExcelExporter._normalize_row(header):
                table_rows = table[1:]
            merged.extend(table_rows)

        return merged

    @staticmethod
    def _normalize_row(row: List[str]) -> List[str]:
        return [str(cell).strip().lower() for cell in row]
